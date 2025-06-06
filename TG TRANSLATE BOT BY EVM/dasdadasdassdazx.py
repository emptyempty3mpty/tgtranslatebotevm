import asyncio
from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command
from deep_translator import GoogleTranslator
from dotenv import load_dotenv
import openpyxl
import os


load_dotenv(dotenv_path='token.env')

API_TOKEN = os.getenv('API_TOKEN')

bot = Bot(token=API_TOKEN)
dp = Dispatcher()

# Состояния для хранения текста
user_texts = {}

# кнопка помощи и предложений
global_keyboard = types.ReplyKeyboardMarkup(
    keyboard=[
        [types.KeyboardButton(text='/help'), types.KeyboardButton(text='/suggestions')]
    ],
    resize_keyboard=True
)


# выбор языка для перевода
def get_language_keyboard():
    buttons = [
        [types.InlineKeyboardButton(text="\U0001f1ec\U0001f1e7Английский", callback_data="lang_en")],
        [types.InlineKeyboardButton(text="\U0001f1f7\U0001f1faРусский", callback_data="lang_ru")],
        [types.InlineKeyboardButton(text="\U0001f1e9\U0001f1eaНемецкий", callback_data="lang_de")],
        [types.InlineKeyboardButton(text="\U0001f1ea\U0001f1f8Испанский", callback_data="lang_es")]
    ]
    return types.InlineKeyboardMarkup(inline_keyboard=buttons)


# start
@dp.message(Command("start"))
async def cmd_start(message: types.Message):
    await message.answer(
        "\U0001f91dПриветствуем Вас в нашем боте. Для получения сведений о нашем боте, нажмите на кнопку /help.",
        reply_markup=global_keyboard
    )


# предложения
@dp.message(Command("suggestions"))
async def cmd_suggestions(message: types.Message):
    await message.answer(
        "Если у вас возникли вопросы или появилось какое-либо предложение, обратитесь к @SEWERO8",
        reply_markup=global_keyboard
    )


# help
@dp.message(Command("help"))
async def cmd_help(message: types.Message):
    help_text = (
        "👋Привет\nЭтот бот является итоговым проектом по дисциплине ЭВМ. Работали над проектом:\n"
        "Чудин Станислав Александрович, Иванов Максим Игоревич, Галенский Аркадий Олегович.\n"
        "Перевод доступен в четырех языках - Английский, Русский, Немецкий и Испанский.\n"
        "Чтобы перевести ваше сообщение:\n"
        "1\uFE0F\u20E3Отправьте сообщение, которое хотите перевести.\n"
        "2\uFE0F\u20E3Выберите нужный язык из предложенных вариантов.\n\n"
        "Если у вас есть предложения по улучшению бота - /suggestions"
    )
    await message.answer(help_text, reply_markup=global_keyboard)


# обработка текста
@dp.message()
async def handle_text(message: types.Message):
    if message.text.startswith('/'):
        return

    # сохранение текста пользователя
    user_texts[message.chat.id] = message.text

    await message.answer(
        "Выбери язык для перевода:",
        reply_markup=get_language_keyboard()
    )


# обработка выбора языка
@dp.callback_query(lambda c: c.data.startswith("lang_"))
async def process_language(callback: types.CallbackQuery):
    lang = callback.data.split("_")[1]
    user_id = callback.from_user.id

    # бот получает сохраненный текст
    text_to_translate = user_texts.get(user_id)

    if not text_to_translate:
        await callback.answer("Текст для перевода не найден", show_alert=True)
        return

    try:
        translated = GoogleTranslator(source='auto', target=lang).translate(text_to_translate)
        await callback.message.edit_text(f"Перевод: {translated}")
        save_to_excel(user_id, translated, text_to_translate)
    except Exception as e:
        await callback.message.edit_text("Ошибка при переводе")
    finally:
        # очистка сохраненного текста
        user_texts.pop(user_id, None)


# сохранение данных в Excel
def save_to_excel(user_id, translated, text_to_translate):
    try:
        wb = openpyxl.load_workbook('BD.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    sheet = wb.active
    next_row = sheet.max_row + 1

    sheet.cell(row=next_row, column=1, value=user_id)
    sheet.cell(row=next_row, column=2, value=translated)
    sheet.cell(row=next_row, column=3, value=text_to_translate)

    wb.save('BD.xlsx')


async def main():
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())